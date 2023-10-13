
try:
    from openpyxl import load_workbook
    from sys import argv
    import pandas as pd
    from datetime import datetime, timedelta
    #import math


    sVersion = "ProcessData v. 2023-09-04.009"




    gYear = '2023'

    sCompanies = ['DI', 'TTE']

    sDataDir = "Data-IN"
    sExcelDir= "Excel"

    dCats={
    'LC':['Lizens'], 
    'MS':['Saas'], 
    'PS':['TTE', 'Online', 'Waffenkartei', 'INDIVIDUAL', 'DIPPS', 'LVS', 'ORKWARE', 'RFID', 'T&T'], 
    'HW':['Handelsware', 'HANDELSWARE']
    }

    sCats = ['Licenses', 'Maintenance Fees', 'Professional Services Fees', 'Hardware']
    sStatus = ['Issued', 'Overdue', '1st Reminder', '2nd Reminder', '3rd Reminder']

    minRevenue = 500

    fTemplate = ''

except:
    print('exception in Init')


#----------------------------------------------------------------------------------------------------------------------------------
#
#
#
#----------------------------------------------------------------------------------------------------------------------------------
def DeleteAllBut(wb, sheetName):
   for s in wb.sheetnames:
      if s != sheetName:
         del wb[s]




#----------------------------------------------------------------------------------------------------------------------------------
#
#
#
#----------------------------------------------------------------------------------------------------------------------------------
def LoadDataFrame(fName: str) -> pd.DataFrame:

    print(fName)
    dfIn = pd.read_csv(fName, sep=';', encoding = 'latin1')
    dfIn.index=Index = dfIn['PEOPLE_ID']

    #print(dfIn.to_string())

    dfOut = dfIn[['PEOPLE_ID', 'KUNDE', 'LAND']].drop_duplicates()
    print(dfOut.to_string())
    for cat in sCats:
        dfOut[cat] = dfIn[dfIn['KLASSE'] == cat].NETTOWERT

    dfOut.fillna(0, inplace = True)

    dfOut['Sum'] = dfOut[sCats].sum(axis = 1)


    dfOut.sort_values(by = ['Sum'], ascending = False, inplace = True, ignore_index=True)
    #print(dfOut.to_string())

    return dfOut




#----------------------------------------------------------------------------------------------------------------------------------
#
#
#
#----------------------------------------------------------------------------------------------------------------------------------
def Process_WonLost(path:str='.', month:str='10', numDetailRows:int=4):
    global sDataDir, sCompanies, minRevenue, gYear, fTemplate

    print('\n  --> Process_WonLost() ...')
    sWonLost = ['Won', 'Lost']

    row = 2  # header of 'WON

    wbTemplate = load_workbook(fTemplate)
    shTemplate = wbTemplate['Won&Lost']
    
    
    for wonLost in sWonLost:
        shTemplate[f'A{row}'].value = f"{wonLost} in {month}'{gYear}"   #Kopfzeile über Tabelle (Bsp. "Won in 10'2022")
        row += 3
        for company in sCompanies:

            fName = f'{path}/{sDataDir}/{wonLost}-{company}.csv'
            
            dfOut = LoadDataFrame(fName)
            dfOut.to_csv(f'Process_WonLost-{wonLost}-{company}.csv', index=None, sep=';', mode='a')
            
            s=sCats+['Sum']
            if company=='DI':
                sumDI=dfOut[s].sum()
            else:
                sumTTE=dfOut[s].sum()

            for i in range(numDetailRows):
                rx = dfOut.loc[i]
                if rx['Sum'] < minRevenue:
                    break;

                shTemplate[f'A{row}'].value = rx['KUNDE']   
                shTemplate[f'B{row}'].value =  round(rx[s[0]])
                shTemplate[f'C{row}'].value =  round(rx[s[1]])
                shTemplate[f'D{row}'].value =  round(rx[s[2]])
                shTemplate[f'E{row}'].value =  round(rx[s[3]])
                shTemplate[f'F{row}'].value =  round(rx[s[4]])
                shTemplate[f'H{row}'].value = company
            
                row += 1

            row += 1 # empty line between DI & TTE
        
        if row <= 15:
            row = 14
        else:
            row = 28

        dsSum = sumDI + sumTTE
        #print(sumDI)
        #print(sumTTE)
        #print(dsSum)

        shTemplate[f'B{row}'].value =  round(dsSum[sCats[0]])
        shTemplate[f'C{row}'].value =  round(dsSum[sCats[1]])
        shTemplate[f'D{row}'].value =  round(dsSum[sCats[2]])
        shTemplate[f'E{row}'].value =  round(dsSum[sCats[3]])
        shTemplate[f'F{row}'].value =  round(dsSum['Sum'])

        row = 16   #Header of 'LOST'
    
    

    DeleteAllBut(wbTemplate, 'Won&Lost')

    wbTemplate.save(f'{path}/{sExcelDir}/Won&Lost.xlsx')




#----------------------------------------------------------------------------------------------------------------------------------
#
#
#
#----------------------------------------------------------------------------------------------------------------------------------
def Process_Offers(path:str='.', month:str='10', numDetailRows:int=10):
    global sDataDir, sCompanies, minRevenue, gYear, fTemplate

    print('\n  --> Process_Offers() ...')

    cntTotal = 0
    sumTotal = 0

    row = 1  # header of 'WON

    wbTemplate = load_workbook(fTemplate)
    shTemplate = wbTemplate['Offers']

    for company in sCompanies:

        fName = f'{path}/{sDataDir}/Angebote-{company}.csv'
            
        dfOut = LoadDataFrame(fName)
        
        dfOut.to_csv(f'Process_Offers-{company}.csv', index=None, sep=';', mode='a')
        
        s=sCats+['Sum']
        if 'sumX'  in locals():
            sumY=sumX.copy()

        sumX=dfOut[s].sum()
        
        row += 1
        
        for i in range(numDetailRows):
            if dfOut.loc[i]['Sum'] < minRevenue:
                break;

            shTemplate[f'A{row}'].value = dfOut.loc[i]['KUNDE']   
            shTemplate[f'B{row}'].value =  round(dfOut.loc[i][s[0]])
            shTemplate[f'C{row}'].value =  round(dfOut.loc[i][s[1]])
            shTemplate[f'D{row}'].value =  round(dfOut.loc[i][s[2]])
            shTemplate[f'E{row}'].value =  round(dfOut.loc[i][s[3]])
            shTemplate[f'F{row}'].value =  round(dfOut.loc[i][s[4]])
            shTemplate[f'G{row}'].value =  dfOut.loc[i]['LAND']
            shTemplate[f'H{row}'].value =  'pending'
            shTemplate[f'I{row}'].value = company
        
            row += 1

        cntTotal += len(dfOut)
        sumTotal += round(sumX[s[4]])
        shTemplate[f'A{row}'].value = '---------'
        shTemplate[f'B{row}'].value =  round(sumX[s[0]])
        shTemplate[f'C{row}'].value =  round(sumX[s[1]])
        shTemplate[f'D{row}'].value =  round(sumX[s[2]])
        shTemplate[f'E{row}'].value =  round(sumX[s[3]])
        shTemplate[f'F{row}'].value =  round(sumX[s[4]])
        shTemplate[f'G{row}'].value =  f'{len(dfOut)-numDetailRows} more offers'
        shTemplate[f'H{row}'].value =  'pending'
        shTemplate[f'I{row}'].value = company
        row += 1 # empty line between DI & TTE
    

    dsSum = sumX + sumY
    #print(dsSum)

    shTemplate[f'B{row}'].value =  round(dsSum[s[0]])
    shTemplate[f'C{row}'].value =  round(dsSum[s[1]])
    shTemplate[f'D{row}'].value =  round(dsSum[s[2]])
    shTemplate[f'E{row}'].value =  round(dsSum[s[3]])
    shTemplate[f'F{row}'].value =  sumTotal
    shTemplate[f'G{row}'].value =  f'{cntTotal} penting offers'
    #shTemplate[f'F{row}'].value =  round(dsSum['Sum'])


    DeleteAllBut(wbTemplate, 'Offers')

    wbTemplate.save(f'{path}/{sExcelDir}/Offers.xlsx')




'''
#----------------------------------------------------------------------------------------------------------------------------------
#
#
# 
#----------------------------------------------------------------------------------------------------------------------------------
def Process_Invoices(path:str='.', month:str='10', numDetailRows:int=10):
    global fTemplate

    print('\n  --> Process_Invoices() ...')

    wbTemplate = load_workbook(fTemplate)
    sheet = wbTemplate['Invoices']
    addDays = 5
    due = datetime.today() - timedelta(days=addDays)
    sDue=due.strftime('%Y-%m-%d 00:00:00')

    row = 5
    for company in sCompanies:

        fName = f'{path}/{sDataDir}/rechnungen-{company}.csv'
        print(fName)
        dfIn = pd.read_csv(fName , sep=';', encoding = 'latin1')
        #print(dfIn.to_string())
        #dfIn.to_csv(f'Process_Invoices-{company}.csv', index=None, sep=';', mode='a')
        
        dStatus = dict.fromkeys(sStatus, 0)


        totalOpen = dfIn.ARK_OPEN.sum()
        dStatus['Issued']  =  dfIn.loc[dfIn['ARK_DUE_DATE'] >= sDue].ARK_OPEN.sum()
        dStatus['1st Reminder'] = dfIn.loc[dfIn['WARNINGS'] == 1].ARK_OPEN.sum()
        dStatus['2nd Reminder'] = dfIn.loc[dfIn['WARNINGS'] == 2].ARK_OPEN.sum()
        dStatus['3rd Reminder'] = dfIn.loc[dfIn['WARNINGS'] == 3].ARK_OPEN.sum()
        sumX = sum(dStatus.values())
        dStatus['Overdue'] = totalOpen - sumX
      
      
        for n in dStatus:
            if dStatus[n] > 0:
                try:
                    #sVal = f'{dStatus[n]:,.2f}'.replace(',', '*').replace('.', ',').replace('*', '.')
                    #print(dStatus[n], sVal)
                    if n == 'Overdue':
                        sheet[f'A{row}'].value = n + f' (more than {addDays} days)'
                    else:
                        sheet[f'A{row}'].value = n 

                    sheet[f'F{row}'].value = float(dStatus[n])
                    row += 1
                except Exception as inst:
                    print(type(inst))    # the exception instance
                    print(inst.args)     # arguments stored in .args
                    print(inst)          # __str__ allows args to be printed directly,
                                        # but may be overridden in exception subclasses
                    x, y = inst.args     # unpack args
                    print('x =', x)
                    print('y =', y)

        print(f'--- total : {totalOpen:,.2f} ---')
        #sheet[f'A{row}'].value = 'total'
        #sheet[f'A{row}'].value = f'{sum(df["Offene Posten"]):,.2f}'

        row = 19

    DeleteAllBut(wbTemplate, 'Invoices')
    wbTemplate.save(f'{path}/{sExcelDir}/Invoices.xlsx')

'''


#----------------------------------------------------------------------------------------------------------------------------------
#
#
# 
#----------------------------------------------------------------------------------------------------------------------------------
def Process_OrdersSum(path:str='.', month:str='10', numDetailRows:int=10):
    global sDataDir, sCompanies, gYear

    print('\n  --> Process_OrdersSum() ...')

    wbTemplate = load_workbook(f'{path}/Excel/Templates.xlsx')
    sheet = wbTemplate['Orders']

    row = 6
    for company in sCompanies:

        fName = f'{path}/{sDataDir}/ka_netto-{company}.csv'
        print(fName)
        dfIn = pd.read_csv(fName , sep=';', encoding = 'latin1')

        dfIn.drop(dfIn[dfIn.AUFTRAG > 'KW'].index, inplace = True)  #remove KW.... orders
        dx = dfIn.groupby('MATCHCODE')['NOCH_NICHT_IN_RECHNUNG_GEST'].sum()
        dx.to_csv(f'Process_OrdersSum-{company}.csv', index=None, sep=';', mode='a')

        sx=0
        dRes={}
        for cat in dCats:
            rev = dx.filter(items = dCats[cat], axis=0).sum()
            sx += rev
            dRes[cat] = rev
            
        others = dx.sum() - sx
        if others > 1:
            dRes['OTHERS'] = others
        else:
            dRes['HW'] += others

        print(dRes)

        try:
            sheet[f'A{row}'].value = 'Backlog'
            sheet[f'B{row}'].value = round(dRes['LC'])
            sheet[f'C{row}'].value = round(dRes['MS'])
            sheet[f'D{row}'].value = round(dRes['PS'])
            sheet[f'E{row}'].value = round(dRes['HW'])
        except:
            pass

        row = 20

    DeleteAllBut(wbTemplate, 'Orders')
    wbTemplate.save(f'{path}/{sExcelDir}/Orders.xlsx')




#----------------------------------------------------------------------------------------------------------------------------------
#
#
# 
#----------------------------------------------------------------------------------------------------------------------------------
def Process_OrdersDetail(path:str='.', sMonth:str='', numDetailRows:int=10):
    global sDataDir, sCompanies, gYear

    print('\n  --> Process_OrdersDetail() ...')

   
    if not sMonth.isnumeric():
       sMonth = '02'   # for debugging only

    if sMonth == '12':
        year = int(gYear) + 1
        month = 1
    else:
        year = int(gYear)
        month = int(sMonth)+1
   
   
    wbTemplate = load_workbook(f'{path}/Excel/Templates.xlsx')
    sheet = wbTemplate['Orders']
    numLines = 8
    row = 5
    for company in sCompanies:
      fName = f'{path}/{sDataDir}/{company}_{year}_{month:02}_KA_Nettowerte_Pos.csv'
      print(fName)
      dfIn = pd.read_csv(fName , sep=';', thousands='.', decimal=',', encoding = 'latin1')
      #print(dfIn.to_string())
      dfIn.drop(dfIn[dfIn.AUFTRAG > 'KW'].index, inplace = True)  #remove KW.... orders

      dfIn['CAT'] = (dfIn['PRODUKTKLASSE'] / 1000).astype(int)
      sQry ='CAT != 81'
      dfIn.query(sQry, inplace = True) 

      cnt = dfIn.shape[0]  # number of rows

      dfIn['WIP_OPEN'] = dfIn['KAP_NETTO'] - dfIn['KAP_NETTO_GEL_MIT_RECHNUNG'] - dfIn['WIP_NETTO']
      dx = dfIn.groupby('CAT')['WIP_OPEN'].sum()  #group summs
      
      dd = dfIn.groupby('KUNDE')['WIP_OPEN'].aggregate(['sum', 'count'])
      
      sCats=[80, 82, 83]
    
      for cat in sCats:
        dd[str(cat)] = dfIn.loc[dfIn['CAT'] == cat].groupby('KUNDE')['WIP_OPEN'].sum()
       
      #dd['Sum'] = dd[sCats].sum(axis = 1)
      dd.fillna(0, inplace=True)
      dd.sort_values(by = ['sum'], ascending = False, inplace = True, ignore_index=False)
    
      dd.reset_index(inplace=True)
      print(dd.to_string())

      
      #print(dfPS.to_string())
      dd.to_csv(f'Process_OrdersDetail-{company}.csv', index=None, sep=';', mode='a')
      dx.to_csv(f'Process_OrdersDetail(Sum)-{company}.csv', index=None, sep=';', mode='a')

      sumPS = 0    #print some detail lines
      for i in range(min(numLines, cnt)):
         try:
            rx = dd.iloc[i]
            kName = rx['KUNDE']
            sheet[f'A{row}'].value = kName
            sheet[f'B{row}'].value = round(rx['80'])
            sheet[f'D{row}'].value = round(rx['82'])
            sheet[f'E{row}'].value = round(rx['83'])
            sheet[f'F{row}'].value = round(rx['sum'])
            #ss = str(dd[dd.KUNDE==kName]['KA_REFERENZ']).split('\n')
            #sheet[f'G{row}'].value = ss[0][3:].strip()[:80]

            row += 1
         except Exception as e:
            print(type(e))    # the exception instance
            print(e)          # __str__ allows args to be printed directly,

      dxTop8 = dd.iloc[0:8].sum()  #group summs  
      print(dxTop8.to_string())

      dRes = {}
      for cat in sCats:
         rev = dx.loc[cat]
         top8 = dxTop8.loc[str(cat)]
         dRes[cat] = rev - top8

      print(dRes)
      
      try:
         sheet[f'A{row}'].value = f'{cnt - numLines} more ...'
         sheet[f'B{row}'].value = round(dRes[80])
         sheet[f'C{row}'].value = 0
         sheet[f'D{row}'].value = round(dRes[82]) - sumPS
         sheet[f'E{row}'].value = round(dRes[83])
      except Exception as inst:
         print(type(inst))    # the exception instance
         print(inst.args)     # arguments stored in .args
         print(inst)          # __str__ allows args to be printed directly,
                              # but may be overridden in exception subclasses
         x, y = inst.args     # unpack args
         print('x =', x)
         print('y =', y)

      row = 19

    DeleteAllBut(wbTemplate, 'Orders')
    wbTemplate.save(f'{path}/{sExcelDir}/Orders.xlsx')





###################################################################################################################################################################################
###################################################################################################################################################################################
if __name__ == '__main__':

    print('#### ', sVersion)
    
    if len(argv) >= 3:
        path = argv[1]
        month = argv[2]
        try:
            gYear = argv[3]
        except:
            pass  #use pre-defined year
    else:
        month = 'xx'
        path = f'test/{month}'
        print("Usage: python won&lost.py <path> <month>")
        #exit(1)

    try:
        fTemplate = f'{path}/{sExcelDir}/Templates.xlsx'
        
        print(fTemplate, gYear, month)

        #Process_Offers(path, month)

        Process_OrdersDetail(path, month)   
        #Process_WonLost(path, month)
        ####Process_Invoices(path, month)  
        

        print('... ProcessData -done- \n\n')

    except Exception as e:
        print("******** ERROR in ProcessData *********")
        print(e)          

        exit(-1)
        

'''

HandleOffers('DI', 10, 0)
HandleOffers('TTE', 10, 12)

sheet[f'G25'].value = f'{totalCnt} pending offers'

wbOffers.save('Offers@Customer.xlsx')

'''



